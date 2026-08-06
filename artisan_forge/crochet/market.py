"""Market research in, a better listing out.

Feed this module a competitor scrape - Etsy search results with tag search
volumes, prices, sales estimates and opportunity scores - and it works out what
to actually put on the listing:

  * which tags to use, ranked by search volume *and* by how well the listings
    using them are performing, filtered to Etsy's 13 x 20-character limit
  * what to charge, from the real price distribution rather than a guess
  * how the winning titles are structured, so the generated title matches the
    shape buyers are already clicking
  * which competitors are worth imitating

Accepted formats: JSON (array or `{"results": [...]}`), JSONL, CSV/TSV and
Excel (.xlsx, needs openpyxl). Field names are matched loosely, so a scrape
from a different tool still lands in the right place.

Nothing here calls an API. It is a deterministic read of the data, which keeps
it free and gives the model a short, high-signal brief instead of a megabyte
of raw JSON.
"""

from __future__ import annotations

import csv
import dataclasses
import io
import json
import re
import statistics
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

from ..packaging import MAX_TAG_LEN, MAX_TAGS, MAX_TITLE_LEN

# Etsy tags cannot contain most punctuation.
_TAG_CLEAN = re.compile(r"[^a-z0-9 \-']+")
# Words too generic to describe a product niche. Deliberately excludes words
# that name or qualify a product in this domain - "top" is a garment, "plus" is
# a size range, "set" and "pack" describe a bundle.
STOP_WORDS = {
    "the", "a", "an", "and", "or", "for", "with", "of", "to", "in", "on", "by",
    "your", "you", "my", "this", "that", "it", "is", "are", "be", "from", "at",
    "more", "new", "best", "all", "any", "x",
}
# Structural title tokens worth reporting as a pattern rather than a keyword.
STRUCTURE_HINTS = (
    "instant download", "digital download", "pdf", "bundle", "pattern",
    "printable", "commercial use", "beginner", "easy", "no sew", "amigurumi",
    "step by step", "english", "us terms", "photo tutorial",
)


@dataclass
class MarketListing:
    """One competitor listing, normalised."""

    title: str = ""
    description: str = ""
    tags: list[str] = field(default_factory=list)
    tag_volumes: dict[str, int] = field(default_factory=dict)
    price: float | None = None
    original_price: float | None = None
    currency: str = ""
    discount_percent: float | None = None
    on_sale: bool = False
    rating: float | None = None
    review_count: int | None = None
    favorites: int | None = None
    favorites_per_day: float | None = None
    estimated_sales: int | None = None
    estimated_revenue: float | None = None
    demand_score: float | None = None
    opportunity_score: float | None = None
    competitive_gap: float | None = None
    momentum_score: float | None = None
    shop_name: str = ""
    shop_total_sales: int | None = None
    days_since_listed: float | None = None
    image_count: int | None = None
    is_digital: bool = False
    bestseller: bool = False
    category: str = ""
    url: str = ""
    query: str = ""

    # ------------------------------------------------------------ derived
    @property
    def performance(self) -> float:
        """A single 0+ number for "is this listing working".

        Blends the signals a scrape usually has. Sales dominate when present;
        otherwise favourites-per-day and review count stand in for them.
        """
        score = 0.0
        if self.estimated_sales:
            score += float(self.estimated_sales)
        if self.favorites_per_day:
            score += float(self.favorites_per_day) * 20
        elif self.favorites:
            score += float(self.favorites) * 0.5
        if self.review_count:
            score += float(self.review_count) * 5
        if self.bestseller:
            score += 200
        return score

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)


@dataclass
class TagInsight:
    """One candidate tag with the evidence behind it."""

    tag: str
    volume: int = 0
    used_by: int = 0
    sales_behind: float = 0.0
    score: float = 0.0
    relevance: float = 1.0

    @property
    def etsy_safe(self) -> bool:
        return bool(self.tag) and len(self.tag) <= MAX_TAG_LEN

    @property
    def on_niche(self) -> bool:
        """False for tags that belong to a different product niche."""
        return self.relevance >= 0.5


@dataclass
class MarketReport:
    """What the scrape says you should do."""

    listings: int = 0
    queries: list[str] = field(default_factory=list)
    tags: list[TagInsight] = field(default_factory=list)
    price_min: float | None = None
    price_max: float | None = None
    price_median: float | None = None
    price_mean: float | None = None
    original_median: float | None = None
    discount_median: float | None = None
    sale_share: float = 0.0
    currency: str = ""
    suggested_price: float | None = None
    suggested_list_price: float | None = None
    title_words: list[tuple[str, int]] = field(default_factory=list)
    title_structures: list[str] = field(default_factory=list)
    title_length_mean: int = 0
    top_performers: list[MarketListing] = field(default_factory=list)
    demand_mean: float | None = None
    opportunity_mean: float | None = None
    image_count_median: int | None = None
    warnings: list[str] = field(default_factory=list)

    def best_tags(self, limit: int = MAX_TAGS, on_niche_only: bool = True) -> list[str]:
        """Etsy-ready tags, highest scoring first.

        Wrong-niche tags are excluded by default. Tagging a cardigan "crochet
        keychain" does not just waste a slot, it pulls in traffic that will not
        convert - a generic "instant download" is strictly better. The caller
        fills any shortfall from its own generic candidates.
        """
        out: list[str] = []
        for insight in self.tags:
            if not insight.etsy_safe:
                continue
            if on_niche_only and not insight.on_niche:
                continue
            if insight.tag not in out:
                out.append(insight.tag)
            if len(out) >= limit:
                break
        return out

    def to_dict(self) -> dict:
        data = dataclasses.asdict(self)
        data["top_performers"] = [
            {"title": l["title"][:100], "price": l["price"],
             "estimated_sales": l["estimated_sales"], "favorites": l["favorites"]}
            for l in data.get("top_performers", [])
        ]
        return data


# ------------------------------------------------------------------- parsing
def _num(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:[.,]\d+)?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return None


def _int(value: object) -> int | None:
    number = _num(value)
    return int(number) if number is not None else None


def _text(value: object, limit: int = 6000) -> str:
    if isinstance(value, (list, tuple)):
        value = " ".join(str(v) for v in value)
    return " ".join(str(value or "").split())[:limit]


def _bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def _tag_list(value: object) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        try:  # a JSON array inside a CSV cell
            decoded = json.loads(value)
            if isinstance(decoded, list):
                value = decoded
        except (json.JSONDecodeError, ValueError):
            value = [p for p in re.split(r"[,;|]", value)]
    if isinstance(value, dict):
        value = list(value)
    if not isinstance(value, (list, tuple)):
        return []
    return [_text(v, 60) for v in value if _text(v, 60)][:40]


def _tag_volumes(value: object) -> dict[str, int]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (json.JSONDecodeError, ValueError):
            return {}
    if not isinstance(value, dict):
        return {}
    out: dict[str, int] = {}
    for key, raw in value.items():
        volume = _int(raw)
        if volume and volume > 0:
            out[_text(key, 60)] = volume
    return out


# Loose field lookup: try each alias in turn.
def _pick(row: dict, *aliases: str):
    lowered = {str(k).strip().lower().replace("_", "").replace(" ", ""): v for k, v in row.items()}
    for alias in aliases:
        key = alias.lower().replace("_", "").replace(" ", "")
        if key in lowered and lowered[key] not in (None, ""):
            return lowered[key]
    return None


def listing_from_row(row: dict) -> MarketListing:
    """Map one scrape record onto a MarketListing."""
    return MarketListing(
        title=_text(_pick(row, "title", "name", "listingtitle"), 400),
        description=_text(_pick(row, "description", "desc", "details"), 6000),
        tags=_tag_list(_pick(row, "tags", "tag", "keywords")),
        tag_volumes=_tag_volumes(_pick(row, "tagVolumes", "tag_volumes", "tagvolume")),
        price=_num(_pick(row, "price", "ehuntPrice", "currentPrice")),
        original_price=_num(_pick(row, "originalPrice", "ehuntOriginalPrice", "listPrice")),
        currency=_text(_pick(row, "currency", "ehuntCurrency", "currencyCode"), 8),
        discount_percent=_num(_pick(row, "ehuntDiscountPercent", "discountPercent", "discount")),
        on_sale=_bool(_pick(row, "onSale", "on_sale")),
        rating=_num(_pick(row, "rating", "ehuntShopRating", "starRating")),
        review_count=_int(_pick(row, "reviewCount", "shopReviewCount", "reviews")),
        favorites=_int(_pick(row, "favoritesCount", "favorites", "numFavorers")),
        favorites_per_day=_num(
            _pick(row, "favoritesPerDayLifetime", "favoritesPerDay", "favoritesDelta")
        ),
        estimated_sales=_int(_pick(row, "ehuntEstimatedSales", "estimatedSales", "sales")),
        estimated_revenue=_num(_pick(row, "ehuntEstimatedRevenue", "estimatedRevenue", "revenue")),
        demand_score=_num(_pick(row, "demandScore", "demand")),
        opportunity_score=_num(_pick(row, "opportunityScore", "opportunity")),
        competitive_gap=_num(_pick(row, "competitiveGapScore", "competitiveGap")),
        momentum_score=_num(_pick(row, "momentumScore", "momentum")),
        shop_name=_text(_pick(row, "shopName", "shop", "seller"), 80),
        shop_total_sales=_int(_pick(row, "shopTotalSales", "shopSales")),
        days_since_listed=_num(_pick(row, "daysSinceListed", "ageDays")),
        image_count=_int(_pick(row, "imageCount", "images")),
        is_digital=_bool(_pick(row, "isDigital", "digital")) or
        _text(_pick(row, "productType"), 20).lower() == "digital",
        bestseller=_bool(_pick(row, "bestseller", "ehuntBestSeller", "isBestSeller")),
        category=_text(_pick(row, "categoryPath", "category"), 120),
        url=_text(_pick(row, "url", "listingUrl", "link"), 300),
        query=_text(_pick(row, "query", "searchTerm", "keyword"), 80),
    )


def _rows_from_json(text: str) -> list[dict] | None:
    try:
        data = json.loads(text)
    except (json.JSONDecodeError, ValueError):
        return None
    if isinstance(data, dict):
        for key in ("results", "listings", "items", "data", "products", "rows"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
        else:
            data = [data]
    if not isinstance(data, list):
        return None
    return [row for row in data if isinstance(row, dict)] or None


def _rows_from_jsonl(text: str) -> list[dict] | None:
    rows: list[dict] = []
    for line in text.splitlines():
        line = line.strip().rstrip(",")
        if not line or line in ("[", "]"):
            continue
        try:
            record = json.loads(line)
        except (json.JSONDecodeError, ValueError):
            continue
        if isinstance(record, dict):
            rows.append(record)
    return rows or None


def _rows_from_csv(text: str) -> list[dict] | None:
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        if "\t" in sample:
            dialect = csv.excel_tab  # type: ignore[assignment]
        elif sample.count(",") >= 2:
            dialect = csv.excel  # type: ignore[assignment]
        else:
            return None
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames or len(reader.fieldnames) < 2:
        return None
    rows = [{k: v for k, v in row.items() if k} for row in reader]
    return [row for row in rows if any(str(v or "").strip() for v in row.values())] or None


def _rows_from_excel(path: Path) -> list[dict]:
    """Read the first worksheet of an .xlsx file."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise ValueError(
            "Reading .xlsx needs openpyxl. Install it with: pip install openpyxl, "
            "or export your data as CSV or JSON instead."
        ) from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        names = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)]
        out: list[dict] = []
        for values in rows:
            if values is None or not any(v is not None and str(v).strip() for v in values):
                continue
            out.append({names[i]: v for i, v in enumerate(values) if i < len(names)})
            if len(out) >= 5000:
                break
        return out
    finally:
        workbook.close()


def parse_market_text(text: str) -> list[MarketListing]:
    """Parse pasted market data in any of the text formats."""
    text = (text or "").strip()
    if not text:
        return []
    for parser in (_rows_from_json, _rows_from_jsonl, _rows_from_csv):
        rows = parser(text)
        if rows:
            return [listing_from_row(row) for row in rows]
    return []


def load_market_data(
    text: str = "",
    files: list[str | Path] | None = None,
) -> tuple[list[MarketListing], list[str]]:
    """Read pasted text plus any uploaded files into one list of competitors."""
    warnings: list[str] = []
    listings: list[MarketListing] = []

    if text.strip():
        found = parse_market_text(text)
        if found:
            listings.extend(found)
        else:
            warnings.append(
                "The pasted market data could not be parsed. Expected JSON, JSONL or CSV."
            )

    for raw in files or []:
        path = Path(raw)
        try:
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                rows = _rows_from_excel(path)
                listings.extend(listing_from_row(row) for row in rows)
            else:
                content = path.read_text(encoding="utf-8", errors="replace")
                found = parse_market_text(content)
                if not found:
                    warnings.append(f"Could not parse {path.name} as JSON, JSONL or CSV.")
                listings.extend(found)
        except Exception as exc:  # noqa: BLE001 - skip the file, keep the run
            warnings.append(f"Could not read {path.name}: {type(exc).__name__}: {exc}")

    usable = [l for l in listings if l.title or l.tags]
    if listings and not usable:
        warnings.append("Market data was read but contained no titles or tags.")
    return usable, warnings


# ------------------------------------------------------------------ analysis
def clean_tag(value: str) -> str:
    """Lowercase, strip punctuation Etsy rejects, collapse spaces."""
    tag = _TAG_CLEAN.sub(" ", str(value or "").lower())
    return " ".join(tag.split())[:MAX_TAG_LEN].strip()


# Craft terms that are relevant to any pattern listing, whatever the item is.
UNIVERSAL_TERMS = {
    "crochet", "pattern", "patterns", "pdf", "digital", "download", "instant",
    "printable", "written", "chart", "diagram", "handmade", "diy", "amigurumi",
    "beginner", "easy", "tutorial", "guide", "yarn", "hook", "stitch",
}
# Digits buried between letters in a long run-together token, e.g.
# "24in1pokemon": keyword-stuffing artefacts whose reported volumes are not
# believable. Short alphanumerics (3xl, 8mm, 4ply) are legitimate and safe.
_INTERIOR_DIGIT = re.compile(r"[a-z]\d+[a-z]")


def looks_like_spam(tag: str) -> bool:
    """True for keyword-stuffed junk tags that scrapers pick up."""
    return any(
        len(token) >= 8 and _INTERIOR_DIGIT.search(token)
        for token in tag.split()
    )


# Garment families: a cardigan listing should still rank "crochet sweater"
# highly, because buyers use the words interchangeably.
GARMENT_FAMILIES: tuple[set[str], ...] = (
    {"cardigan", "cardi", "sweater", "jumper", "pullover", "hoodie"},
    {"top", "tee", "tank", "camisole", "crop", "blouse", "vest"},
    {"dress", "pinafore", "romper", "jumpsuit"},
    {"blanket", "afghan", "throw", "quilt"},
    {"hat", "beanie", "bonnet", "cloche", "cap"},
    {"bag", "tote", "purse", "pouch", "basket", "backpack"},
    {"scarf", "cowl", "shawl", "wrap", "poncho", "snood"},
    {"amigurumi", "plush", "plushie", "toy", "doll", "stuffie", "keychain"},
    {"socks", "slippers", "booties", "slipper"},
    {"mittens", "gloves", "mitts"},
    {"coaster", "placemat", "doily", "potholder", "dishcloth", "washcloth"},
)
# Broader categories. Within a category, items are close enough that a tag is
# still worth carrying (a cardigan listing can reasonably say "crochet top").
# Across categories it is not: a cardigan tagged "crochet keychain" attracts
# traffic that will never convert.
CATEGORIES: dict[str, set[str]] = {
    "garment": {
        "cardigan", "cardi", "sweater", "jumper", "pullover", "hoodie", "top",
        "tee", "tank", "camisole", "crop", "blouse", "vest", "dress",
        "pinafore", "romper", "jumpsuit", "skirt", "shrug", "bolero",
    },
    "accessory": {
        "hat", "beanie", "bonnet", "cloche", "cap", "scarf", "cowl", "shawl",
        "wrap", "poncho", "snood", "mittens", "gloves", "mitts", "socks",
        "slippers", "booties", "slipper", "headband", "earwarmer",
    },
    "bag": {"bag", "tote", "purse", "pouch", "basket", "backpack", "clutch"},
    "home": {
        "blanket", "afghan", "throw", "quilt", "coaster", "placemat", "doily",
        "potholder", "dishcloth", "washcloth", "rug", "cushion", "pillow",
        "runner", "curtain",
    },
    "toy": {
        "amigurumi", "plush", "plushie", "toy", "doll", "stuffie", "keychain",
        "keyring", "rattle", "mobile",
    },
}
# Any product noun at all.
PRODUCT_NOUNS: set[str] = {word for words in CATEGORIES.values() for word in words}


def _categories_of(tokens: set[str]) -> set[str]:
    """Which product categories a set of words belongs to."""
    return {name for name, words in CATEGORIES.items() if tokens & words}


def _relevance_tokens(text: str) -> set[str]:
    tokens = {t for t in re.findall(r"[a-z0-9]+", str(text or "").lower()) if len(t) > 2}
    return {t for t in tokens if t not in STOP_WORDS}


def _expand_families(tokens: set[str]) -> set[str]:
    """Add family synonyms for any garment word in the set."""
    expanded = set(tokens)
    for family in GARMENT_FAMILIES:
        if tokens & family:
            expanded |= family
    return expanded


def tag_relevance(tag: str, product_tokens: set[str]) -> float:
    """How on-subject a tag is for this product, from 0.35 to 1.0.

    The distinctive part of a tag is what matters. "crochet pattern" is generic
    and always fine; "amigurumi pokemon" is only fine if you are making a
    Pokemon. Adjacent product nouns sit in between, because a cardigan listing
    can reasonably carry "crochet top" but not "crochet keychain".
    """
    if not product_tokens:
        return 1.0
    distinctive = _relevance_tokens(tag) - UNIVERSAL_TERMS
    if not distinctive:
        return 1.0                              # generic craft tag
    if distinctive & product_tokens:
        return 1.0                              # names this product

    tag_categories = _categories_of(distinctive)
    product_categories = _categories_of(product_tokens)
    if tag_categories and product_categories:
        # same category -> a near neighbour; different category -> wrong shopper
        return 0.7 if (tag_categories & product_categories) else 0.3
    if distinctive <= PRODUCT_NOUNS:
        return 0.3                              # a product we are not making
    return 0.45                                 # an unrecognised descriptor


def _rank_tags(
    listings: list[MarketListing],
    relevance: str = "",
) -> list[TagInsight]:
    """Score every tag seen, by search volume and by who is winning with it.

    Three corrections matter here, because the raw numbers mislead:

    * Volume is scaled logarithmically. A scrape will contain a tag claiming 55M
      searches next to a real one claiming 260K; linear scaling lets the outlier
      swamp everything.
    * Off-topic tags are damped. A "crochet" search returns cardigans next to
      Pokemon keychains, and tagging a cardigan "amigurumi pokemon" helps nobody.
      Pass `relevance` (the item type, yarn, sizes) to keep the list on subject.
    * Keyword-stuffed junk is dropped outright.
    """
    import math

    volumes: dict[str, int] = {}
    used_by: Counter[str] = Counter()
    sales_behind: dict[str, float] = {}

    for listing in listings:
        performance = listing.performance
        # tag_volumes keys and tags may differ in case/spacing
        for raw, volume in listing.tag_volumes.items():
            tag = clean_tag(raw)
            if tag:
                volumes[tag] = max(volumes.get(tag, 0), int(volume))
        for raw in listing.tags:
            tag = clean_tag(raw)
            if not tag:
                continue
            used_by[tag] += 1
            sales_behind[tag] = sales_behind.get(tag, 0.0) + performance

    product_tokens = _expand_families(_relevance_tokens(relevance))
    insights: list[TagInsight] = []
    max_log_volume = math.log1p(max(volumes.values())) if volumes else 0.0
    max_sales = max(sales_behind.values()) if sales_behind else 0.0

    for tag in set(list(volumes) + list(used_by)):
        if looks_like_spam(tag):
            continue
        volume = volumes.get(tag, 0)
        performance = sales_behind.get(tag, 0.0)

        volume_part = (math.log1p(volume) / max_log_volume) if max_log_volume else 0.0
        sales_part = (performance / max_sales) if max_sales else 0.0
        # A tag every competitor uses is more contested, so damp it slightly.
        crowding = 1.0 / (1.0 + 0.12 * max(used_by.get(tag, 1) - 1, 0))
        # On-subject tags win; off-subject ones stay on the list but rank lower.
        relevance_part = tag_relevance(tag, product_tokens)

        score = (0.55 * volume_part + 0.45 * sales_part) * crowding * relevance_part
        insights.append(TagInsight(
            tag=tag, volume=volume, used_by=used_by.get(tag, 0),
            sales_behind=round(performance, 1), score=round(score, 5),
            relevance=relevance_part,
        ))

    insights.sort(key=lambda i: (-i.score, -i.volume, i.tag))
    return insights


def _title_analysis(listings: list[MarketListing]) -> tuple[list[tuple[str, int]], list[str], int]:
    words: Counter[str] = Counter()
    structures: Counter[str] = Counter()
    lengths: list[int] = []

    for listing in listings:
        title = listing.title
        if not title:
            continue
        lengths.append(len(title))
        lowered = title.lower()
        for token in re.findall(r"[a-z0-9']+", lowered):
            if len(token) > 2 and token not in STOP_WORDS and not token.isdigit():
                words[token] += 1
        for hint in STRUCTURE_HINTS:
            if hint in lowered:
                structures[hint] += 1
        if re.search(r"\d{2,}\s*\+", title):
            structures["numeric count claim (e.g. 200+)"] += 1
        if title.count("|") >= 2:
            structures["pipe-separated keyword blocks"] += 1
        if title.count(",") >= 2:
            structures["comma-separated keyword list"] += 1

    mean_length = int(statistics.fmean(lengths)) if lengths else 0
    top_structures = [f"{name} ({count}x)" for name, count in structures.most_common(8)]
    return words.most_common(24), top_structures, mean_length


def _price_stats(listings: list[MarketListing]) -> dict:
    prices = [l.price for l in listings if l.price and l.price > 0]
    originals = [l.original_price for l in listings if l.original_price and l.original_price > 0]
    discounts = [l.discount_percent for l in listings if l.discount_percent]
    on_sale = [l for l in listings if l.on_sale or (l.discount_percent or 0) > 0]
    currencies = Counter(l.currency for l in listings if l.currency)

    stats: dict = {
        "price_min": min(prices) if prices else None,
        "price_max": max(prices) if prices else None,
        "price_median": round(statistics.median(prices), 2) if prices else None,
        "price_mean": round(statistics.fmean(prices), 2) if prices else None,
        "original_median": round(statistics.median(originals), 2) if originals else None,
        "discount_median": round(statistics.median(discounts), 1) if discounts else None,
        "sale_share": round(len(on_sale) / len(listings), 2) if listings else 0.0,
        "currency": currencies.most_common(1)[0][0] if currencies else "",
    }

    # Price where the winners sit, not the whole field: undercutting the bottom
    # of a market full of 75%-off bundles is a race to nothing.
    winners = sorted(listings, key=lambda l: -l.performance)[: max(3, len(listings) // 4)]
    winner_prices = [l.price for l in winners if l.price and l.price > 0]
    if winner_prices:
        stats["suggested_price"] = round(statistics.median(winner_prices), 2)
    elif stats["price_median"]:
        stats["suggested_price"] = stats["price_median"]
    else:
        stats["suggested_price"] = None

    # If the market runs permanent sales, mirror it: list high, discount to the
    # suggested price, so the listing shows a strikethrough like the rest.
    if stats["suggested_price"] and stats["discount_median"] and stats["sale_share"] >= 0.4:
        keep = max(0.15, 1.0 - stats["discount_median"] / 100.0)
        stats["suggested_list_price"] = round(stats["suggested_price"] / keep, 2)
    else:
        stats["suggested_list_price"] = None
    return stats


def analyse(
    listings: list[MarketListing],
    warnings: list[str] | None = None,
    relevance: str = "",
) -> MarketReport:
    """Turn a list of competitor listings into an actionable report.

    `relevance` describes the product being listed (item type, yarn weight,
    sizes). It keeps the tag ranking on subject when the scrape covers a broad
    search term that returned several different niches.
    """
    listings = [l for l in listings if l.title or l.tags]
    report = MarketReport(listings=len(listings), warnings=list(warnings or []))
    if not listings:
        return report

    report.queries = [q for q, _ in Counter(l.query for l in listings if l.query).most_common(5)]
    report.tags = _rank_tags(listings, relevance)
    report.title_words, report.title_structures, report.title_length_mean = _title_analysis(listings)

    for key, value in _price_stats(listings).items():
        setattr(report, key, value)

    report.top_performers = sorted(listings, key=lambda l: -l.performance)[:5]

    demand = [l.demand_score for l in listings if l.demand_score is not None]
    opportunity = [l.opportunity_score for l in listings if l.opportunity_score is not None]
    images = [l.image_count for l in listings if l.image_count]
    report.demand_mean = round(statistics.fmean(demand), 1) if demand else None
    report.opportunity_mean = round(statistics.fmean(opportunity), 1) if opportunity else None
    report.image_count_median = int(statistics.median(images)) if images else None
    return report


# ------------------------------------------------------------------- prompts
def market_brief(report: MarketReport, limit: int = 3000) -> str:
    """Render the report as compact, prompt-ready text."""
    if not report.listings:
        return ""
    lines: list[str] = [f"MARKET RESEARCH ({report.listings} competing listings analysed)"]
    add = lines.append

    if report.queries:
        add(f"Search terms scraped: {', '.join(report.queries)}")

    if report.price_median is not None:
        currency = report.currency or "USD"
        add(
            f"Prices ({currency}): min {report.price_min:g}, median "
            f"{report.price_median:g}, max {report.price_max:g}"
        )
        if report.sale_share:
            add(f"{int(report.sale_share * 100)}% of listings are on sale")
        if report.discount_median:
            add(f"Median discount: {report.discount_median:g}%")
        if report.suggested_price:
            add(f"Price the best performers sit at: {report.suggested_price:g} {currency}")
        if report.suggested_list_price:
            add(
                f"Suggested strategy: list at {report.suggested_list_price:g} and discount to "
                f"{report.suggested_price:g}, matching the market's sale pattern"
            )

    if report.demand_mean is not None:
        add(f"Mean demand score: {report.demand_mean:g}")
    if report.opportunity_mean is not None:
        add(f"Mean opportunity score: {report.opportunity_mean:g}")
    if report.image_count_median:
        add(f"Competitors use a median of {report.image_count_median} listing images")

    usable = [i for i in report.tags if i.etsy_safe][:20]
    if usable:
        add("")
        add("HIGHEST VALUE TAGS (tag | monthly search volume | competitors using it):")
        for insight in usable:
            volume = f"{insight.volume:,}" if insight.volume else "unknown"
            add(f"  {insight.tag} | {volume} | {insight.used_by}")

    oversized = [i.tag for i in report.tags[:30] if not i.etsy_safe]
    if oversized:
        add(
            "High-volume phrases too long for a tag (use them in the title instead): "
            + ", ".join(oversized[:6])
        )

    if report.title_words:
        add("")
        add(
            "Most common title keywords: "
            + ", ".join(f"{word} ({count})" for word, count in report.title_words[:16])
        )
    if report.title_structures:
        add("Title conventions in this niche: " + "; ".join(report.title_structures))
    if report.title_length_mean:
        add(f"Mean competitor title length: {report.title_length_mean} characters")

    if report.top_performers:
        add("")
        add("BEST PERFORMING COMPETITORS:")
        for listing in report.top_performers:
            bits = [f'"{listing.title[:90]}"']
            if listing.price:
                bits.append(f"{listing.price:g} {listing.currency or ''}".strip())
            if listing.estimated_sales:
                bits.append(f"~{listing.estimated_sales} sales")
            if listing.favorites:
                bits.append(f"{listing.favorites} favourites")
            add("  - " + " | ".join(bits))

    return "\n".join(lines)[:limit]


def listing_prompt(
    pattern: dict,
    report: MarketReport,
    garment: str = "",
    shop: str = "",
    source_title: str = "",
) -> str:
    """Ask ChatGPT for the listing, given the pattern and the market data."""
    sizes = (pattern.get("sizes") or {}).get("labels") or []
    currency = report.currency or "USD"
    price_line = (
        f"The best performers sell at about {report.suggested_price:g} {currency}."
        if report.suggested_price else
        "No competitor pricing was available; price it for a premium graded pattern."
    )
    # The source title often carries the character or design name which is the
    # single most important search keyword for buyers looking for this specific
    # pattern. Without it the listing is generic.
    keyword_line = ""
    if source_title:
        keyword_line = (
            f"- IMPORTANT: The pattern is called \"{source_title}\". This name (or the key "
            "words from it) MUST appear in the title and at least two of the tags, because "
            "buyers search for it by name.\n"
        )
    return (
        "You are writing the Etsy listing for a crochet PATTERN PDF (not a finished item). "
        "Use the market research to choose wording and tags that will actually get found.\n\n"
        f"PRODUCT: {pattern.get('title', 'Crochet pattern')}\n"
        f"ITEM TYPE: {garment or 'crochet piece'}\n"
        f"SIZES: {', '.join(sizes) if sizes else 'one size'}\n"
        f"SKILL LEVEL: {pattern.get('skill_level', 'advanced beginner')}\n"
        f"PAGES: it is a complete graded pattern with technical diagrams\n"
        f"SHOP: {shop or 'an independent designer'}\n\n"
        f"{market_brief(report)}\n\n"
        "RULES\n"
        f"- Title: at most {MAX_TITLE_LEN} characters. Lead with the highest-volume "
        "keywords a buyer would actually type. Follow the title conventions above.\n"
        + keyword_line +
        f"- Tags: exactly {MAX_TAGS} tags, each at most {MAX_TAG_LEN} characters, "
        "lowercase, no punctuation. Prefer the high-volume tags listed above, but do "
        "not repeat the same phrase twice, and include a couple of lower-competition "
        "long-tail tags.\n"
        "- Description: open with two lines that sell the outcome, then WHAT YOU GET, "
        "SKILL LEVEL, WHAT'S INCLUDED, HOW IT WORKS and TERMS sections. Mention that "
        "this is a digital download and that the pattern file may not be resold.\n"
        f"- {price_line}\n"
        "- Be honest. Do not claim a bundle size or feature the pattern does not have.\n"
        "- No emoji.\n\n"
        "Return JSON only:\n"
        "{\n"
        '  "title": "...",\n'
        '  "tags": ["13 tags"],\n'
        '  "description": "...",\n'
        '  "suggested_price_usd": 7.5,\n'
        '  "list_price_usd": 12.0,\n'
        '  "materials": ["PDF", "Digital Download"],\n'
        '  "reasoning": "one sentence on why these keywords"\n'
        "}"
    )


def normalise_listing(raw: object, report: MarketReport, fallback: dict) -> dict:
    """Make a model listing response safe, and enforce Etsy's limits."""
    out = dict(fallback)
    if not isinstance(raw, dict):
        return out

    title = _text(raw.get("title"), MAX_TITLE_LEN * 2).strip()
    if len(title) >= 10:
        out["title"] = title[:MAX_TITLE_LEN].rstrip(" |,-")

    tags: list[str] = []
    for candidate in list(raw.get("tags") or []):
        tag = clean_tag(candidate)
        if tag and tag not in tags:
            tags.append(tag)
        if len(tags) >= MAX_TAGS:
            break
    # top up from the research rather than shipping a half-empty tag set
    for tag in report.best_tags(MAX_TAGS):
        if len(tags) >= MAX_TAGS:
            break
        if tag not in tags:
            tags.append(tag)
    if tags:
        out["tags"] = tags

    description = _text(raw.get("description"), 12000)
    if len(description) >= 120:
        out["description"] = str(raw.get("description")).strip()

    price = _num(raw.get("suggested_price_usd"))
    if price and 0.5 <= price <= 200:
        out["suggested_price_usd"] = round(price, 2)
    elif report.suggested_price:
        out["suggested_price_usd"] = report.suggested_price

    list_price = _num(raw.get("list_price_usd"))
    if list_price and list_price > out.get("suggested_price_usd", 0):
        out["list_price_usd"] = round(list_price, 2)
    elif report.suggested_list_price:
        out["list_price_usd"] = report.suggested_list_price

    materials = [_text(m, 40) for m in (raw.get("materials") or []) if _text(m, 40)]
    if materials:
        out["materials"] = materials[:13]

    reasoning = _text(raw.get("reasoning"), 400)
    if reasoning:
        out["keyword_reasoning"] = reasoning
    return out
